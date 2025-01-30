from flask import Flask, request, render_template, send_file, jsonify
import openpyxl
from curl_cffi import requests
from bs4 import BeautifulSoup
import html2text
import os
from threading import Thread
from datetime import datetime
from openpyxl.styles import Font, PatternFill




app = Flask(__name__)

# Настройки для парсинга
BASE_URL = "https://5.188.139.66:443/product-search/api/query"
HEADERS = {
    'Host': 'www.garshinka.ru',
    'accept': 'application/json, text/plain, */*',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
}

converter = html2text.HTML2Text()
converter.ignore_links = True
converter.ignore_images = True
converter.ignore_emphasis = True

# Хранилище для статуса
status = {"processed": 0, "total": 0, "completed": False, "output_file": None}


def fetch_product_data(product_name):
    params = {'search': product_name, 'size': '24', 'offset': '0', 'description': True}
    response = requests.get(BASE_URL, headers=HEADERS, params=params, verify=False)
    data = response.json()

    if not data.get('products'):
        return None, None, None, None, None

    product = data['products'][0]
    link = f"https://www.garshinka.ru{product.get('url', '')}"
    product_name = product.get('title', '')
    product_link = f"https://5.188.139.66:443{product.get('url', '')}"
    product_id = data['products'][0].get('id', '')
    images = data['products'][0].get('images', '')
    image_links = []
    for img in images:
        image = img.replace(f"{img.split('/')[-2]}", "productbig")
        print(image)
        image_links.append(image)

    param_response = requests.get(f'https://5.188.139.66:443/product/get-parameters/{product_id}', headers=HEADERS, verify=False)
    data_params = param_response.json()['parameters']
    spec = []
    for p in data_params:
        k_param = p.get('name', '')
        v_param = p.get('value', '')
        v_param = ', '.join(v_param)  # Преобразуем список в строку
        spec.append({k_param: v_param})
        print(k_param, ": ", v_param)

    product_response = requests.get(product_link, headers=HEADERS, verify=False)
    soup = BeautifulSoup(product_response.text, 'lxml')

    desc = soup.find('div', class_='text-formatted')
    description_text = converter.handle(str(desc)).replace('#', '') if desc else ''

    return product_name, description_text.strip(), '; '.join(image_links), spec, link

# Цвет заливки (бледно-оливковый)
fill_color = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

# Столбцы для окраски
highlight_columns = [4, 5, 6, 9, 22]

def process_excel(file_path, output_path):
    global status
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    total_rows = sum(
        1 for row in sheet.iter_rows(min_row=2)  # Пропускаем заголовок (min_row=2)
        if any(cell.value is not None for cell in row)  # Учитываем только строки с непустыми ячейками
    )
    status["total"] = total_rows
    status["processed"] = 0

    # Окрашиваем заголовки столбцов 5, 6, 9 и 22
    for col_idx in highlight_columns:
        sheet.cell(row=1, column=col_idx).fill = fill_color

    # Обработка данных
    for row in range(3, sheet.max_row + 1):
        product_name = sheet.cell(row=row, column=4).value
        if product_name:
            # Функция fetch_product_data возвращает нужные данные
            name, description, images, spec, link = fetch_product_data(product_name)

            # Записываем основные данные
            sheet.cell(row=row, column=5).value = name
            sheet.cell(row=row, column=6).value = link
            sheet.cell(row=row, column=9).value = description
            sheet.cell(row=row, column=10).value = " "
            sheet.cell(row=row, column=22).value = images
            sheet.cell(row=row, column=23).value = " "

            # Работа со spec
            if spec:  # Проверяем, что spec не пустой
                start_column = 25  # Начальный столбец для spec
                for entry in spec:  # Проходим по вложенным словарям
                    for key, value in entry.items():
                        # Найти текущий столбец для ключа или создать новый
                        col_index = find_or_create_column(sheet, start_column, key)
                        # Записать значение в соответствующую ячейку
                        sheet.cell(row=row, column=col_index).value = ', '.join(value) if isinstance(value,
                                                                                                     list) else value

        status["processed"] += 1

    wb.save(output_path)
    status["completed"] = True  # Указываем, что процесс еще не завершен
    status["processed"] = total_rows  # Завершение


def find_or_create_column(sheet, start_column, key):
    """
    Найти столбец с заголовком key или создать новый, начиная с start_column.
    """
    # Проверяем только столбцы, начиная с start_column
    for col in range(start_column, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header == key:  # Если нашли заголовок
            return col

    # Если столбец не найден, находим первый пустой столбец начиная с start_column
    col = start_column
    while sheet.cell(row=1, column=col).value:  # Пропускаем заполненные столбцы
        col += 1

    # sheet.cell(row=1, column=col).value = key  # Устанавливаем заголовок
    header_cell = sheet.cell(row=1, column=col)
    header_cell.value = key
    header_cell.font = Font(bold=True)
    return col



@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            input_path = "uploaded_file.xlsx"
            output_path = f"result_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            file.save(input_path)

            # Запуск обработки в отдельном потоке
            status["output_file"] = output_path

            thread = Thread(target=process_excel, args=(input_path, output_path))
            thread.start()
            return jsonify({"status": "processing"})

    return render_template('index.html')


@app.route('/status', methods=['GET'])
def get_status():
    return jsonify(status)


@app.route('/download', methods=['GET'])
def download():
    output_path = status.get("output_file")  # Получаем имя файла из статуса
    if not status.get("completed", False):  # Проверяем, завершена ли обработка
        return "Файл ещё не готов", 404
    if os.path.exists(output_path):
        return send_file(output_path, as_attachment=True)
    return "Файл не найден", 404


if __name__ == '__main__':
    app.run(debug=True)